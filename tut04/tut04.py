import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook

folder1 = "output_by_subject"

if not os.path.exists(folder1):
    os.makedirs(folder1)
folder2 = "output_individual_roll"

if not os.path.exists(folder2):
    os.makedirs(folder2)


def output_by_subject():
    head = ['rollno', 'register_sem', 'subno', 'sub_type']
    with open('regtable_old.csv', 'r') as info:
        sinfo = csv.reader(info)
        for words in sinfo:
            data = []
            data.append(words[0])
            data.append(words[1])
            data.append(words[3])
            data.append(words[-1])
            if (data[2] == "subno"): continue
            subno = '{}.xlsx'.format(data[2])
            path = './output_by_subject/' + subno

            if (os.path.isfile(path)):
                book = load_workbook('output_by_subject\\{}.xlsx'.format(data[2]))
                sheet = book.active
                sheet.append(data)
                book.save('output_by_subject\\{}.xlsx'.format(data[2]))

            else:
                book = Workbook()
                sheet = book.active
                sheet.append(head)
                sheet.append(data)
                book.save('output_by_subject\\',data[2],'.xlsx')
    return


def output_individual_roll():
    head = ['rollno', 'register_sem', 'subno', 'sub_type']
    with open('regtable_old.csv', 'r') as info:
        sinfo = csv.reader(info)
        for words in sinfo:
            data = []
            data.append(words[0])
            data.append(words[1])
            data.append(words[3])
            data.append(words[-1])
            if (data[0] == "rollno"): continue
            rollno = '{}.xlsx'.format(data[0])
            path = './output_individual_roll/' + rollno
            if (os.path.isfile(path)):  # adding data to sheet
                book = load_workbook('output_individual_roll\\{}.xlsx'.format(data[0]))
                sheet = book.active
                sheet.append(data)
                book.save('output_individual_roll\\{}.xlsx'.format(data[0]))


            else:
                book = Workbook()
                sheet = book.active
                sheet.append(head)
                sheet.append(data)
                book.save('output_individual_roll\\',data[0],'.xlsx')
    return


output_by_subject()
output_individual_roll()