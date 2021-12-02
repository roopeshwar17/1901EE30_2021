import os
import csv
from numpy import NAN,nan
from PIL.Image import linear_gradient
import pandas as pd
from openpyxl import Workbook
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font,Alignment,Border,Side, alignment
import mini

path=os.getcwd()


response_file=pd.read_csv("./sample_input/responses.csv")
master_file=pd.read_csv("./sample_input/master_roll.csv")

sample_output_path = "./sample_output/marksheet"
pos=5
neg=1
zero=0
final_score=[]
status=[]



def generate_marksheet(pos=5,neg=1):
    os.makedirs("sample_output",exist_ok = True)
    os.makedirs(sample_output_path,exist_ok = True)
    if(neg <1):
        
        neg=-neg

    print(master_file.head())
    master_roll=master_file['roll'].values.tolist()
    master_name=master_file['name'].values.tolist()


    dict_roll={}
    for i in range(len(master_roll)):
        dict_roll[master_roll[i]]=master_name[i]
        print(dict_roll[master_roll[i]])

    
    print(response_file.head())
    response_roll=response_file['Roll Number'].values.tolist()
    correct_options=response_file.iloc[0].values.tolist()
    correct_options=correct_options[6:]

   
    no_stu_master=0
    no_of_response=0
    os.chdir(path)

    alignment_heading=Alignment(horizontal='right',vertical='bottom')
    alignment_content=Alignment(horizontal='left',vertical='bottom')
    alignment_ans=Alignment(horizontal='center',vertical='bottom')
    font_heading=Font(name='Calibri',size=14,bold=False)
    font_content=Font(name='Calibri',size=14,bold=True)
    right_color=Font(color='00FF00',name='Calibri',size=14,bold=False)
    left_color=Font(color='ff0000',name='Calibri',size=14,bold=False)
    give_color=Font(color='0000FF',name='Calibri',size=14,bold=False)
    border_style=Side(border_style='medium',color='000000')

    while(no_stu_master<len(master_roll)):
        wb=Workbook()
        sheet=wb.active
        img=Image('iitp_logo.png')
        
        border=Border(top=border_style,bottom=border_style,left=border_style,right=border_style)
        img.width=610
        img.height=80
        sheet.add_image(img,'A1')
        sheet.column_dimensions['A'].width=17
        sheet.column_dimensions['B'].width=17
        sheet.column_dimensions['C'].width=17
        sheet.column_dimensions['D'].width=17
        sheet.column_dimensions['E'].width=17

        sheet['A6']='Name :'
        sheet['A6'].font=font_heading
        sheet['A6'].alignment=alignment_heading
        
        sheet['D6']='Exam :'
        sheet['D6'].font=font_heading
        sheet['D6'].alignment=alignment_heading
        sheet['E6']='quiz'
        sheet['E6'].font=font_content
        sheet['E6'].alignment=alignment_content
        sheet['A7']='Roll Number :'
        sheet['A7'].font=font_heading
        sheet['A7'].alignment=alignment_heading

        empty=['A9','A10','A11','A12','B9','C9','D9','E9','A15','B15','D15','E15']
        for key in empty:
            sheet[key].font=font_content
            sheet[key].alignment=alignment_ans
            sheet[key].border=border
        
        mini.function(sheet,alignment_ans,font_content,alignment_content,dict_roll,master_roll,no_stu_master,correct_options,pos,neg,zero,border)
        
        line=16
        m=0
        while(m<len(correct_options)):
            if(line<=40):
                sheet.cell(row=line,column=2).font=Font(size=12,color='000000FF', name='Calibri')
                sheet.cell(row=line,column=2).alignment=Alignment('center')
                sheet.cell(row=line,column=2).value=correct_options[m]
                sheet.cell(row=line,column=2).border=border

            else:
                sheet.cell(row=line-25,column=5).font=Font(size=12,color='000000FF', name='Calibri')
                sheet.cell(row=line-25,column=5).alignment=Alignment('center')
                sheet.cell(row=line-25,column=5).value=correct_options[m]
                sheet.cell(row=line-25,column=5).border=border

            line+=1
            m+=1

       

        sheet.cell(row=10,column=1).alignment=Alignment('center')
        sheet.cell(row=10,column=1).font=Font(size=12,bold=True,name="Calibri")
        sheet.cell(row=10,column=1).value='No.'

        sheet.cell(row=11,column=1).alignment=Alignment('center')
        sheet.cell(row=11,column=1).font=Font(size=12,bold=True,name="Calibri")
        sheet.cell(row=11,column=1).value='Marking'

        sheet.cell(row=12,column=1).alignment=Alignment('center')
        sheet.cell(row=12,column=1).font=Font(size=12,bold=True,name="Calibri")
        sheet.cell(row=12,column=1).value='Total'

        if(no_of_response<len(response_roll) and master_roll[no_stu_master]==response_roll[no_of_response]):

            stu_options=[]
            stu_options=response_file.iloc[no_of_response].values.tolist()
            stu_options=stu_options[6:]

            k=0
            total=0
            wrong=0
            right=0
            notattempt=0
            while(k<len(correct_options)):
                if(stu_options[k]==correct_options[k]):
                    right+=1
                k+=1

            nonattempt=stu_options.count(nan)
            wrong=len(stu_options)-right-notattempt
            status.append('[{},{},{}]'.format(right,wrong,notattempt))

            total=((right*pos)-(wrong*neg))
            max_marks=len(correct_options)*pos
            final_score.append('{}/{}'.format(total,max_marks))

            

            sheet.cell(row=10,column=2).font=Font(size=12,color="00008000",name="Calibri")
            sheet.cell(row=10,column=2).alignment=Alignment('center')
            sheet.cell(row=10,column=2).value=str(right)
            sheet.cell(row=10,column=2).border=border

            sheet.cell(row=10,column=3).font=Font(size=12,color="00FF0000",name="Calibri")
            sheet.cell(row=10,column=3).alignment=Alignment('center')
            sheet.cell(row=10,column=3).value=str(wrong)
            sheet.cell(row=10,column=3).border=border

            sheet.cell(row=10,column=4).font=Font(size=12,name="Calibri")
            sheet.cell(row=10,column=4).alignment=Alignment('center')
            sheet.cell(row=10,column=4).value=str(notattempt)
            sheet.cell(row=10,column=4).border=border
            
            sheet.cell(row=12,column=2).font=Font(size=12,color="00008000",name="Calibri")
            sheet.cell(row=12,column=2).alignment=Alignment('center')
            sheet.cell(row=12,column=2).value=str(right*pos)
            sheet.cell(row=12,column=2).border=border

            sheet.cell(row=12,column=3).font=Font(size=12,color="00FF0000",name="Calibri")
            sheet.cell(row=12,column=3).alignment=Alignment('center')
            sheet.cell(row=12,column=3).value=str(wrong*neg)
            sheet.cell(row=12,column=3).border=border

            sheet.cell(row=12,column=5).font=Font(size=12,name="Calibri",color='000000FF')
            sheet.cell(row=12,column=5).alignment=Alignment('center')
            sheet.cell(row=12,column=5).value='{}/{}'.format(total,max_marks)
            sheet.cell(row=12,column=5).border=border

            line=16
            m=0
            while(m<len(correct_options)):
                if(line<=40):
                    if(stu_options[m]==correct_options[m]):
                        sheet.cell(row=line,column=1).font=Font(size=12,color="00008000",name="Calibri")
                        sheet.cell(row=line,column=1).alignment=Alignment('center')
                        sheet.cell(row=line,column=1).value=stu_options[m]
                        sheet.cell(row=line,column=1).border=border

                    else:
                        sheet.cell(row=line,column=1).font=Font(size=12,color="00FF0000",name="Calibri")
                        sheet.cell(row=line,column=1).alignment=Alignment('center')
                        sheet.cell(row=line,column=1).value=stu_options[m]
                        sheet.cell(row=line,column=1).border=border

                else:
                    if(stu_options[m]==correct_options[m]):
                        sheet.cell(row=line-25,column=4).font=Font(size=12,color="00008000",name="Calibri")
                        sheet.cell(row=line-25,column=4).alignment=Alignment('center')
                        sheet.cell(row=line-25,column=4).border=border
                        sheet.cell(row=line-25,column=4).value=stu_options[m]
                    else:
                        sheet.cell(row=line-25,column=4).font=Font(size=12,color="00FF0000",name="Calibri")
                        sheet.cell(row=line-25,column=4).border=border
                        sheet.cell(row=line-25,column=4).alignment=Alignment('center')
                        sheet.cell(row=line-25,column=4).value=stu_options[m]
                line+=1
                m+=1
            no_of_response+=1
        
        wb.save(f'./sample_output/marksheet/{master_roll[no_stu_master]}.xlsx')
        no_stu_master+=1

def generate_concise_marksheet():
    response_file.insert(5,'score_after_negative',final_score,True)
    response_file['Status']=status
    response_file.to_excel('./sample_output/marksheet/concise_marksheet.xlsx')

def get_roll_email():
    lst = []
    for index,row in response_file.iterrows():
        lst.append([row["Roll Number"],row["Email address"],row["IITP webmail"]])
    return lst
        








    
    





