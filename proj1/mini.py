import os
import csv
from numpy import NAN,nan
from PIL.Image import linear_gradient
import pandas as pd
from openpyxl import Workbook
import openpyxl
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
from openpyxl.styles import Font,Alignment,Border,Side, alignment


def function(sheet,alignment_ans,font_content,alignment_content,dict_roll,master_roll,no_stu_master,correct_options,pos,neg,zero,border):       
        lst=['B11','C11','E11','D11','D12','E10']
        for key in lst:
            sheet[key].border=border


        sheet.merge_cells('A5:E5')
        sheet.cell(row=5,column=1).font=Font(size=18,bold=True,name='Century',underline='single')
        sheet.cell(row=5,column=1).alignment=alignment_ans
        sheet.cell(row=5,column=1).value='Mark Sheet'

        sheet.merge_cells('B6:C6')
        sheet.cell(row=6,column=2).font=font_content
        sheet.cell(row=6,column=2).alignment=alignment_content
        sheet.cell(row=6,column=2).value=dict_roll[master_roll[no_stu_master]]

        sheet.cell(row=7,column=2).font=font_content
        sheet.cell(row=7,column=2).alignment=alignment_content
        sheet.cell(row=7,column=2).value=master_roll[no_stu_master]

        sheet.cell(row=10,column=5).font=font_content
        sheet.cell(row=10,column=5).alignment=alignment_ans
        sheet.cell(row=10,column=5).value=str(len(correct_options))

        sheet.cell(row=9,column=2).font=Font(size=12,bold=True, name='Calibri')
        sheet.cell(row=9,column=2).alignment=Alignment('center')
        sheet.cell(row=9,column=2).value='Right'

        sheet.cell(row=9,column=3).font=Font(size=12,bold=True, name='Calibri')
        sheet.cell(row=9,column=3).alignment=Alignment('center')
        sheet.cell(row=9,column=3).value='Wrong'

        sheet.cell(row=9,column=4).font=Font(size=12,bold=True, name='Calibri')
        sheet.cell(row=9,column=4).alignment=Alignment('center')
        sheet.cell(row=9,column=4).value='Not Attempt'

        sheet.cell(row=9,column=5).font=Font(size=12,bold=True, name='Calibri')
        sheet.cell(row=9,column=5).alignment=Alignment('center')
        sheet.cell(row=9,column=5).value='Max'

        sheet.cell(row=11,column=2).font=Font(size=12,color='00008000', name='Calibri')
        sheet.cell(row=11,column=2).alignment=Alignment('center')
        sheet.cell(row=11,column=2).value=str(pos)

        sheet.cell(row=11,column=3).font=Font(size=12,color='00FF0000', name='Calibri')
        sheet.cell(row=11,column=3).alignment=Alignment('center')
        sheet.cell(row=11,column=3).value=str(neg)

        sheet.cell(row=11,column=4).font=Font(size=12, name='Calibri')
        sheet.cell(row=11,column=4).alignment=Alignment('center')
        sheet.cell(row=11,column=4).value=str(zero)

        sheet.cell(row=15,column=1).font=Font(size=12,bold=True, name='Calibri')
        sheet.cell(row=15,column=1).alignment=Alignment('center')
        sheet.cell(row=15,column=1).value='Student Ans'

        sheet.cell(row=15,column=4).font=Font(size=12,bold=True, name='Calibri')
        sheet.cell(row=15,column=4).alignment=Alignment('center')
        sheet.cell(row=15,column=4).value='Student Ans'

        sheet.cell(row=15,column=2).font=Font(size=12,bold=True, name='Calibri')
        sheet.cell(row=15,column=2).alignment=Alignment('center')
        sheet.cell(row=15,column=2).value='Correct Ans'

        sheet.cell(row=15,column=5).font=Font(size=12,bold=True, name='Calibri')
        sheet.cell(row=15,column=5).alignment=Alignment('center')
        sheet.cell(row=15,column=5).value='Correct Ans'
