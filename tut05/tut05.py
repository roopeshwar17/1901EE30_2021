import os
from openpyxl import Workbook
from openpyxl import load_workbook
import csv
def create_workbook(record):
    wb=Workbook()
    ws=wb.active
    ws.title = "Overall"
    wb.save(f'output\\{record[0]}.xlsx')
    return

def sheet_isexist(sheet_name,wb):
    for sheet in wb.sheetnames:
        if sheet == sheet_name:
            return 1
    return 0

def get_credits_spi(wb,credit_map):
    Spi,Credits= [],[]
    for sheet in wb.sheetnames[1:]:
        ws = wb[sheet]
        credits = [int(cell.value) for cell in ws["E"][1:]]
        spi = [credit_map[cell.value.strip().strip("*")] for cell in ws["G"][1:]]
        Credits.append(sum(credits))
        Spi.append(round(sum([spi[i]*credits[i] for i in range(len(spi))])/sum(credits),2))
    return Spi,Credits

def get_results(Spi,ws,Credits):        
    ws.append(["Semester No"]+[x for x in range(1,9)])
    ws.append(["Semester wise Credit taken"]+Credits)
    ws.append(["SPI"]+Spi)
    ws.append(["Total Credits taken"]+[sum(Credits[:i+1]) for i in range(len(Credits))])
    ws.append(["CPI"]+[round(sum(Spi[:i+1])/(i+1),2) for i in range(len(Spi))])
    return
def generate_marksheet():
    dir_name =  "output"
    if not os.path.exists(dir_name):
        os.mkdir(dir_name)
    course_data = open("subjects_master.csv", "r")
    course_csv=csv.reader(course_data)
    course_list =  [list(record) for record in course_csv][1:] 
    course_dict = {}
    max_length = 0
    for record in course_list:
        subno = record[0]
        if subno not in course_dict:
            course_dict[subno] = record[1:]
            max_length = max(max_length,len(record[1]))  
    grades_data = open("grades.csv", "r")
    grades_csv = csv.reader(grades_data)
    grades_list = [list(record) for record in grades_csv][1:] 
    c=1
    for record in grades_list:
        Roll,Sem_no, = record[0],record[1]
        file_path='./output/'+'{}.xlsx'.format(Roll)
        if not os.path.isfile(file_path):
            create_workbook(record) 
        w_b=load_workbook(r'output\\{}.xlsx'.format(Roll))
        if not sheet_isexist(f'Sem{Sem_no}',w_b):
            w_b.create_sheet(f'Sem{Sem_no}',int(Sem_no))
            w_s = w_b[f"Sem{int(Sem_no)}"]
            w_s.column_dimensions["C"].width = max_length
        w_s = w_b[f"Sem{int(Sem_no)}"]
        print(f"{c}Creating {Sem_no}")
        if w_s.max_row==1 :
            w_s.append(["Sl No.","Subject No.","Subject Name","L-T-P","Credit","Subject Type","Grade"])
        SubCode,Credit,Grade,Sub_Type = record[2:]
        subname,ltp,crd = course_dict[f"{SubCode}"][0],course_dict[f"{SubCode}"][1],course_dict[f"{SubCode}"][2]
        w_s.append([w_s.max_row,SubCode,subname,ltp,crd,Sub_Type,Grade])
        w_b.save(r'output\\{}.xlsx'.format(Roll))
        c+=1;       
    credit_map = {'AA':10,'AB':9,'BB':8,'BC':7,'CC':6,'CD':5,'DD':4,'F':0,'I':0,
                'AA*':10,'AB*':9,'BB*':8,'BC*':7,'CC*':6,'CD*':5,'DD*':4,'F*':0,'I*':0}
    names_data = open("names-roll.csv","r")
    names_csv = csv.reader(names_data)
    names_list = [list(record) for record in names_csv][1:] 
    for record in names_list:
        name,Roll = record[1],record[0]
        w_b = load_workbook(r'output\\{}.xlsx'.format(Roll))
        w_s =w_b.active
        w_s.column_dimensions["A"].width = 30
        w_s.append(["RollNo",Roll])
        w_s.append(["Name of Student",name])
        w_s.append(["Discipline",Roll[4:6]])
        Spi,Credits= get_credits_spi(w_b,credit_map)
        get_results(Spi,w_s,Credits)
        print(Spi)
        w_b.save(r'output\\{}.xlsx'.format(Roll))
    return
generate_marksheet()