import os
import csv
from openpyxl import Workbook
from openpyxl import load_workbook

print('compiling please wait')

def helper(sub_dict_ltp,studinfo_dict,final_result):
	for roll_no in registered_course:
		for row in registered_course[roll_no]:
			codesub = row[2]
			temp = sub_dict_ltp[codesub]
			for x in temp:
				check = False
				if roll_no in students_feedback:
					for row1 in students_feedback[roll_no]:
						if row1[1] == codesub:
							if int(row1[2]) == int(x):
								check = True
								break
				if check == False:
					if roll_no not in studinfo_dict:
						temp2 = [roll_no,row[0],row[1],codesub,'NA_IN_STUDENTINFO','NA_IN_STUDENTINFO','NA_IN_STUDENTINFO','NA_IN_STUDENTINFO']
						final_result.append(temp2)
						break
					else:
						temp1 = [roll_no,row[0],row[1],codesub,studinfo_dict[roll_no][3],studinfo_dict[roll_no][0],studinfo_dict[roll_no][1],studinfo_dict[roll_no][2]]
						final_result.append(temp1)
						
						break

def excel_maker(ans,filename):
	if os.path.isfile(filename):
		wb = load_workbook(r'course_feedback_remaining.xlsx')
	else:
		wb = Workbook()
	sheet = wb.active
	lst = ['rollno','register_sem','schedule_sem','subno','Name','email','aemail','contact']
	sheet.append(lst)
	for ro in ans:
		sheet.append(ro)
	wb.save(filename)

def dict_2dlist_maker(filename,diction,list_append,var_roll):
		file= open(filename,'r')  
		reader = csv.DictReader(file)
		for row in reader:
			if row[var_roll] in diction:
				
				diction[row[var_roll]].append([row[list_append[0]],row[list_append[1]],row[list_append[2]]])
			else:
				
				diction[row[var_roll]] = []
				diction[row[var_roll]].append([row[list_append[0]],row[list_append[1]],row[list_append[2]]])
		file.close()
		return diction

registered_course = {}
registered_course=dict_2dlist_maker('course_registered_by_all_students.csv',registered_course,['register_sem','schedule_sem','subno'],'rollno')
sub_dict_ltp={}
studinfo_dict = {}
students_feedback = {}
students_feedback = dict_2dlist_maker('course_feedback_submitted_by_students.csv',students_feedback,['stud_name','course_code','feedback_type'],'stud_roll')

def feedback_not_submitted():

	ltp_mapping_feedback_type = {1: 'lecture', 2: 'tutorial', 3:'practical'}
	output_file_name = "course_feedback_remaining.xlsx"


	
	file= open('course_master_dont_open_in_excel.csv','r') 
	reader = csv.DictReader(file)
	for row in reader:
		lst = row['ltp'].split('-')
		lst1 =[]
		for el in range(len(lst)):
			if lst[el] !='0':
				s=el+1
				lst1.append(s)
				
		sub_dict_ltp[row['subno']] = lst1
	file.close()
	
	
	file= open('studentinfo.csv','r')
	reader = csv.DictReader(file)
	for row in reader:
		studinfo_dict[row['Roll No']] = [row['email'],row['aemail'],row['contact'],row['Name']]
	file.close()

	final_result = []
	helper(sub_dict_ltp,studinfo_dict,final_result)
					

	excel_maker(final_result,output_file_name)

feedback_not_submitted()

print("completed successfully")